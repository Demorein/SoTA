#Class with basic kernel functions
class SoTACore():

    #Read json file. Returns a list
    #EXAMPLE: print(_json_read(file)["Variable"])
    def _json_read(file:str) -> str:
        import json
        with open(file, "r") as f:
            return(json.load(f))
        
    #Write Log in file, Returns None
    #EXAMPLE: _log("Hello", log)
    def _logs(log:str, file = "log") -> None:
        with open(f"{file}.txt", "a") as f:
            f.write(f"{log}")

    #Returns a list with date and time
    #EXAMPLE1: _time()[0] --> return(time)
    #EXAMPLE2: _time()[0] --> return(date)
    def _time() -> str:
        import datetime
        dt = datetime.datetime.now()
        nowtime = dt.time()
        nowdate = dt.date()
        return([f"{nowtime.hour}:{nowtime.minute}:{nowtime.second}",f"{nowdate.day}.{nowdate.month}.{nowdate.year}"])
    
    def _check(object:int, xlsx:str) -> bool:
        row = 1
        import openpyxl
        wb = openpyxl.load_workbook(xlsx)
        ws = wb.active
        while True:
            row += 1
            if ws[f"A{row}"].value == None:
                wb.close()
                return(False)
            elif id == int(ws[f"A{row}"].value):
                wb.close()
                return(True)
            else:
                pass
    
    #Class for managing server via discord
    class Server():
        #returns bool
        #Starts .bat the Minecraft server. Accepts the startup file directory
        def _start(platform:str) -> bool:
            from os import path, system
            if path.exists(f"{platform}.exe") == True:
                system(f"start {platform}.exe")
                return(True)
            else:
                return(False)
            
        def _close(platform:str) -> bool:
            from os import path
            from subprocess import Popen
            if path.exists(f"{platform}.exe") == True:
                Popen(f'taskkill /im {platform}SoTA /f')
                return(True)
            else:
                return(False)